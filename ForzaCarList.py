from bs4 import BeautifulSoup
from collections import defaultdict
import pandas as pd
import requests
import openpyxl as pxl


# The purpose of this function is to compile a list of manufacturers in Forza
# @input: string variable which holds the link of all manufacturers in Forza
#    > should be 'https://forza.fandom.com/wiki/Category:Manufacturers'
# @output: an array full of manufacturers in the game
def getListOfManufacturers(link):

    arr = []

    print("Aquiring list of manufacturers in Forza...")

    # Sends a GET request to the provided URL and stores the response in the 'page' variable
    page = requests.get(link)

    print("...")

    # Parses the retrieved HTML content from the page using BeautifulSoup
    # The 'html.parser' argument specifies that we're dealing with an HTML document
    soup = BeautifulSoup(page.content, 'html.parser')

    # Searches for a specific <div> element with the class 'category-page__members'
    # This <div> is assumed to contain a list of manufacturers
    list_items = soup.find('div', class_='category-page__members') 

    # Check if the 'list_items' was found (i.e., the <div> exists on the page)
    if list_items:
        # Finds all <li> (list item) elements within the <div>, assuming each manufacturer is in a list format
        list_items = list_items.find_all('li')

        # Loop through each <li> element
        for item in list_items:
            # Extracts the text content of each list item and strips any leading/trailing whitespace
            # The manufacturer name is stored in the text of the <li> element
            element = (item.text).strip()

            # Rename Briggs Automotive to BAC
            if element == 'Briggs Automotive Company':
                element  = 'BAC'

            # Drop '(manufacturer) from Lego Speed Champions
            if element ==  'LEGO Speed Champions (manufacturer)':
                element = 'LEGO Speed Champions'

            # There's a mistake in the page where 'Category:Manufacturers By Origin' is listed as a manufacturer
            # Dumps that value and any other unnecessary value.
            if element != 'Category:Manufacturers By Origin':
                    arr.append(element)
    
    print("List of manufacturers aquired.")
  
    return arr


# The purpose of the function is to strip data from the Forza website for all car models
# @input: links -> an array that holds all the links it should comb over.
# @output: an array that has all the car models stored in it.
def getListOfCars(link):

    cars = []

    for url in links:
        # Sends a GET request to the URL to retrieve the web page content
        page = requests.get(url)

        print(f"Scraping \'{url}\'...")
        
        # Parses the retrieved page content using BeautifulSoup and creates a BeautifulSoup object
        # The 'html.parser' argument specifies that we are parsing an HTML document
        parsed_page = BeautifulSoup(page.content, 'html.parser')
        
        # Selects all table rows (<tr>) that are inside a table body (<tbody>)
        # This assumes that the car data is stored within rows of a table on the webpage
        table_rows = parsed_page.select('tbody tr')
        
        # Loop through each row found in the table body
        for row in table_rows:
            # Select the specific elements inside the row that match 'td div a' (anchors within divs inside table cells)
            # It is assumed that the car data is stored in these anchor tags
            data = row.select('td:first-child div a')

            # If we found any matching data within the row
            if data:
                try:
                    # Extracts the text from the first anchor tag and combines it with the text of its next sibling element (if present)
                    # This concatenation assumes that car details might be split across elements
                    car = data[0].text + ' ' + data[0].next_sibling.strip()
                except AttributeError:
                    # If an AttributeError occurs (likely because the next_sibling doesn't exist or is of an unexpected type),
                    # print an error message for debugging purposes
                    print("Throwing attribute error. Idk why...")

                if car not in cars:
                    cars.append(car)
        print("...")

    print("Finished aquiring all cars.")

    cars = sorted(cars, key=lambda car: car[-4:], reverse=True)
    return cars


# This function simply matches the models with a manufacturer.
# @input: array of manufacturers, array of models
# @output: filled out excel sheet
def insertModels(manufacturers, models):
    # Initialize the workbook as None
    workbook = None

    # Cycle through the list of manufacturers
    for manufacturer in manufacturers:
        
        print(f"Finding all models made by {manufacturer}.")

        # Initialize an empty temporary array that holds all the models for a certain car manufacturer
        temp_arr = []

        # Cycle through the models array
        index = 0

        while index < len(models):

            model = models[index]

            # Sees if the element has the manufacturer.
            if model.startswith(manufacturer):
                
                # If it matches the manufacturer, then append it to temp_arr
                model = model.removeprefix(manufacturer).strip()
                temp_arr.append(model)

                # Pop 'model' out of the 'models' array. The reasoning behind this is so the models array 
                # becomes smaller meaning its less time cosuming.
                models.pop(index)
                print('*pop*')

            else:
                # Increment the index only if no pop was performed
                index += 1

        # At the end of this cycle, it should get all the models with the manufacturer name.
        # Now we sort the elements in the array by descending order.
        temp_arr = sorted(temp_arr, key=lambda car: car[-4:], reverse = True)

        # Pass in the current list of models and the current manufacturer
        workbook = insertIntoExcel(temp_arr, manufacturer, workbook)
    
    # Save the final workbook after processing all manufacturers
    if workbook:
        workbook.save("Forza Car List.xlsx")
        workbook.close()

# Inserts models into an excel sheet if the data doesn't exist in the excel sheet yet. Also, if no excel sheet it passed
#       into the function, then it will create an empty excel sheet with the proper headers.
# @input: The current array of models to be inserted, the current manufacturer who's car models we are working with,
#           and the workbook that we are working with.             
# @output: Should return an excel sheet that has all the data we need. (Workbook object)
def insertIntoExcel(temp_arr, manufacturer, excel_workbook = None):
    
    if excel_workbook is None:
        # Create a new workbook and select an active sheet
        wb = pxl.Workbook()
        ws = wb.active

        # Set coloumn width 
        ws.column_dimensions['A'].width = 16.22  # Make
        ws.column_dimensions['B'].width = 56.77  # Model
        ws.column_dimensions['C'].width = 35.55  # Paint Name
        ws.column_dimensions['D'].width = 16.22  # Paint Type
        
        # Create header for chart
        # Color the headers black
        for row in ws["A1:J2"]:
            for cell in row:
                cell.fill = pxl.styles.PatternFill(start_color = 'FF000000', end_color = 'FF000000', fill_type = 'solid')

        # Create the title for each coloumn
        ws['A2'] = 'Make'
        ws['B2'] = 'Model'
        ws['C2'] = 'Color Name'
        ws['D2'] = 'Paint Type'
        
        ws.merge_cells('E1:G1')
        ws['E1'].font = pxl.styles.Font(color = 'FF2F75B5') # Set color 1's color to blue
        ws['E1'].alignment = pxl.styles.Alignment(horizontal="center")
        ws['E1'] = 'COLOR 1 (X)'
        ws['E2'] = 'Hue'
        ws['F2'] = 'Saturation'
        ws['G2'] = 'Brightness'

        ws.merge_cells('H1:J1')
        ws['H1'].font = pxl.styles.Font(color = 'FFFFD966') # Set color 2's color to yellow
        ws['H1'].alignment = pxl.styles.Alignment(horizontal="center")
        ws['H1'] = 'COLOR 2 (Y)'
        ws['H2'] = 'Hue'
        ws['I2'] = 'Saturation'
        ws['J2'] = 'Brightness'

        # Set the color for the text in the header cells to white
        for cell in ws.iter_rows(min_row = 2, max_row = 2, min_col = 1, max_col = 10):
            for c in cell:
                c.font = pxl.styles.Font(color = 'FFFFFFFF')

    else:
        # Use the existing workbook
        wb = excel_workbook
        ws = wb.active

    # Fill in the array
    # Find the first empty row
    first_empty_row = ws.max_row + 1

    # Append the models from the first empty row
    for row in range(len(temp_arr)):
        ws[f'A{first_empty_row + row}'] = manufacturer
        ws[f'B{first_empty_row + row}'] = temp_arr[row]

    # Save the workbook
    wb.save("Forza Car List.xlsx")

    # Close the workbook
    wb.close()

    return wb
            
        
if __name__ == '__main__':
    link_for_manufacturers = 'https://forza.fandom.com/wiki/Category:Manufacturers'
    links = ["https://forza.fandom.com/wiki/Forza_Horizon_3/Cars",
            "https://forza.fandom.com/wiki/Forza_Horizon_4/Cars",
            "https://forza.fandom.com/wiki/Forza_Horizon_5/Cars"]

    manufacturers = getListOfManufacturers(link_for_manufacturers)

    models = getListOfCars(links)

    manufacturers.sort(key = lambda x: x.lower())

    insertModels(manufacturers, models)